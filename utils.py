import json
import os
import requests
from tempfile import gettempdir
import time
import csv
from openpyxl import Workbook, load_workbook

from config import Configuration


class Utility:

    #Reading from a file
    @staticmethod
    def readFromFile(fileName):
        try:
            with open(fileName, "r") as f:
                my_dict = json.load(f)
                return my_dict
        except FileNotFoundError:
            print("File - "+fileName+" not found")

    @staticmethod
    def readFromFile1(fileName,noOfRecords):
        try:
            with open(fileName, "r") as f:
                my_dict = json.load(f)
                retValue = {}
                for symbol in my_dict:
                    listOfValues = (my_dict[symbol])[0:noOfRecords]
                    my_dict[symbol] = listOfValues
                    print("Inside readFromFile with noOfRecords = ",len(listOfValues))

                return my_dict
        except FileNotFoundError:
            print("File - "+fileName+" not found")

    
    @staticmethod
    def read_from_file(file_name):
        try:
            with open(file_name, 'r') as f:
                return json.load(f)
        except FileNotFoundError:
            print(f"File not found: {file_name}")
            return None

    #Writing to a File
    @staticmethod
    def writeToFile(fileName, data):
        with open(fileName, "w") as f:
            json.dump(data, f)

    @staticmethod
    def writeToCSVFile(fileName, data):
        with open(fileName, "w", newline='') as file:
            writer = csv.writer(file)
            writer.writerows(data)

    @staticmethod
    def writeToExcelFolder(folder, fileName, sheetName, data):
        os.makedirs(folder, exist_ok=True)
        Utility.writeToExcel(folder+fileName, sheetName, data)

    @staticmethod
    def writeToExcel(fileName, sheetName, data):
        
        # Check if file exists
        if os.path.exists(fileName):
            # If exists, load the workbook
            wb = load_workbook(fileName)
            print(f"Opened existing file: {fileName}")

        else:
            # If not exists, create a new workbook
            wb = Workbook()
            print(f"Created new file: {fileName}")
        
        # Get list of all sheet names
        sheets = wb.sheetnames

        # Check if a sheet exists

        if sheetName in sheets:
            sheet = wb[sheetName]
        else:
            sheet = wb.create_sheet(title=sheetName)    
        # Write data starting at row 1
        for row_idx, row_data in enumerate(data, start=1):
            for col_idx, value in enumerate(row_data, start=1):
                sheet.cell(row=row_idx, column=col_idx, value=value)
        wb.save(fileName)

    @staticmethod
    def sentGetRequest( url, headers, requestParams):

        # Make a GET request with headers and parameters
        response = requests.get(url, headers=headers, params=requestParams)
        
        print("======================")

        return response.text

    @staticmethod
    def findMinMax(array):
        min = 0
        max = 0
        if len(array) > 0:
            min = array[0]
            max = array[0]
            for ele in array:
                if min > ele:
                    min = ele
                if max < ele:
                    max = ele
        return min, max

    @staticmethod
    def getMinMaxValue(candles):
        
        
        # Initialize min to first candle Low and max to first candle high 
        minVal = candles[0][3]
        maxVal = candles[0][2]
        minCloseVal = candles[0][4]
        maxCloseVal = candles[0][4]
        
        for candle in candles:
            if minVal > candle[3]:
                minVal = candle[3]
            
            if maxVal < candle[2]:
                maxVal = candle[2]


            

        return minVal, maxVal

    @staticmethod
    def getDiffWithMinMax(candles):
        diffWithMin = []
        diffWithMax = []

        minVal, maxVal = Utility.getMinMaxValue(candles)
        for candle in candles :
            diffWithMin.append(round(candle[4] - minVal,2))
            diffWithMax.append(round(maxVal - candle[4],2))

        return diffWithMin,diffWithMax

    @staticmethod
    def findTrend(candles, remarks):
        length = len(candles)
        finalClose = candles[0][4]
        closeDiffArr = []
        for i in range(1, length - 1):
            close = round(finalClose - candles[i][4],2)
            closeDiffArr.append(close)

        #print(remarks,":::::",closeDiffArr)
        return closeDiffArr

    @staticmethod
    def getTableForTimeFrame(timeframe: str) -> str:
        # Search the array Configuration.timeframe_tablemap for the timeframe to table map
        for tf, table in Configuration.timeframe_tablemap:
            if tf == timeframe:
                return table
        return None
