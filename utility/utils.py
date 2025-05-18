import json
import math
import os
from datetime import timedelta, date, datetime
from zoneinfo import ZoneInfo


import pandas as pd
import requests
from tempfile import gettempdir
import time
import csv
from openpyxl import Workbook, load_workbook
import numpy as np
from service.support_resistance import SupportResistance

from utility.config import Configuration


class Utility:

    #Reading from a file
    @staticmethod
    def readFromFile(fileName):
        try:
            with open(fileName, "r") as f:
                my_dict = json.load(f)
                return my_dict
        except FileNotFoundError:
            print("File - "+fileName+" not found")

    @staticmethod
    def readFromFile1(fileName,noOfRecords):
        try:
            with open(fileName, "r") as f:
                my_dict = json.load(f)
                retValue = {}
                for symbol in my_dict:
                    listOfValues = (my_dict[symbol])[0:noOfRecords]
                    my_dict[symbol] = listOfValues
                    print("Inside readFromFile with noOfRecords = ",len(listOfValues))

                return my_dict
        except FileNotFoundError:
            print("File - "+fileName+" not found")

    
    @staticmethod
    def read_from_file(file_name):
        try:
            with open(file_name, 'r') as f:
                return json.load(f)
        except FileNotFoundError:
            print(f"File not found: {file_name}")
            return None

    #Writing to a File
    @staticmethod
    def writeToFile(fileName, data):
        with open(fileName, "w") as f:
            json.dump(data, f)

    @staticmethod
    def writeToCSVFile(fileName, data):
        with open(fileName, "w", newline='') as file:
            writer = csv.writer(file)
            writer.writerows(data)

    @staticmethod
    def writeToExcelFolder(folder, fileName, sheetName, data):
        os.makedirs(folder, exist_ok=True)
        Utility.writeToExcel(folder+fileName, sheetName, data)

    @staticmethod
    def writeToExcel(fileName, sheetName, data):
        
        # Check if file exists
        if os.path.exists(fileName):
            # If exists, load the workbook
            wb = load_workbook(fileName)
            print(f"Opened existing file: {fileName}")

        else:
            # If not exists, create a new workbook
            wb = Workbook()
            print(f"Created new file: {fileName}")
        
        # Get list of all sheet names
        sheets = wb.sheetnames

        # Check if a sheet exists

        if sheetName in sheets:
            sheet = wb[sheetName]
        else:
            sheet = wb.create_sheet(title=sheetName)    
        # Write data starting at row 1
        for row_idx, row_data in enumerate(data, start=1):
            for col_idx, value in enumerate(row_data, start=1):
                sheet.cell(row=row_idx, column=col_idx, value=value)
        wb.save(fileName)

    @staticmethod
    def write_dataframe_to_excel(dfs_with_sheets: dict, file_path: str):
        import pandas as pd
        import os
        from openpyxl import load_workbook

        # Check if the Excel file exists
        if os.path.exists(file_path):
            # Load the existing workbook
            book = load_workbook(file_path)

            # Delete only the sheets that are being updated
            for sheet_name in dfs_with_sheets.keys():
                if sheet_name in book.sheetnames:
                    del book[sheet_name]

            # Save workbook after deleting target sheets
            book.save(file_path)

            # Open in append mode to write updated sheets
            with pd.ExcelWriter(file_path, engine='openpyxl', mode='a') as writer:
                for sheet_name, df in dfs_with_sheets.items():
                    if not sheet_name or not isinstance(sheet_name, str):
                        raise ValueError(f"Invalid sheet name: {sheet_name}")
                    df.to_excel(writer, sheet_name=sheet_name[:31], index=False)

        else:
            # File does not exist — create new workbook
            with pd.ExcelWriter(file_path, engine='openpyxl', mode='w') as writer:
                for sheet_name, df in dfs_with_sheets.items():
                    if not sheet_name or not isinstance(sheet_name, str):
                        raise ValueError(f"Invalid sheet name: {sheet_name}")
                    df.to_excel(writer, sheet_name=sheet_name[:31], index=False)

    @staticmethod
    def sentGetRequest( url, headers, requestParams = None):

        # Make a GET request with headers and parameters
        response = requests.get(url, headers=headers, params=requestParams)
        
        # print("====================== Req Params == "+str(requestParams))

        return response.text

    @staticmethod
    def findMinMax(array):
        min = 0
        max = 0
        if len(array) > 0:
            min = array[0]
            max = array[0]
            for ele in array:
                if min > ele:
                    min = ele
                if max < ele:
                    max = ele
        return min, max

    @staticmethod
    def getMinMaxValue(candles):
        
        
        # Initialize min to first candle Low and max to first candle high 
        minVal = candles[0][3]
        maxVal = candles[0][2]
        minCloseVal = candles[0][4]
        maxCloseVal = candles[0][4]
        
        for candle in candles:
            if minVal > candle[3]:
                minVal = candle[3]
            
            if maxVal < candle[2]:
                maxVal = candle[2]


            

        return minVal, maxVal

    @staticmethod
    def getDiffWithMinMax(candles):
        diffWithMin = []
        diffWithMax = []

        minVal, maxVal = Utility.getMinMaxValue(candles)
        for candle in candles :
            diffWithMin.append(round(candle[4] - minVal,2))
            diffWithMax.append(round(maxVal - candle[4],2))

        return diffWithMin,diffWithMax

    @staticmethod
    def findTrend(candles, remarks):
        length = len(candles)
        finalClose = candles[0][4]
        closeDiffArr = []
        for i in range(1, length - 1):
            close = round(finalClose - candles[i][4],2)
            closeDiffArr.append(close)

        #print(remarks,":::::",closeDiffArr)
        return closeDiffArr

    @staticmethod
    def getTableForTimeFrame(timeframe: str) -> str:
        # Search the array Configuration.timeframe_tablemap for the timeframe to table map
        for tf, table in Configuration.all_timeframes_tablemap:
            if tf == timeframe:
                return table
        return None

    @staticmethod
    def get_start_end_date_for_historic_data_fetch(timeframe, no_of_candles_required):
        startDate = None
        endDate = None

        if timeframe == "minutes/1":
            total_candles_in_a_day = 375
        elif timeframe == "minutes/5":
            total_candles_in_a_day = 75
        elif timeframe == "minutes/15":
            total_candles_in_a_day = 25
        elif timeframe == "hours/1":
            total_candles_in_a_day = 7
        elif timeframe == "minutes/75":
            total_candles_in_a_day = 6
        elif timeframe == "days/1":
            total_candles_in_a_day = 1
        elif timeframe == "weeks/1":
            total_candles_in_a_day = 1/4
        elif timeframe == "months/1":
            total_candles_in_a_day = 1/7
        no_days_required = math.ceil(no_of_candles_required/total_candles_in_a_day)+3

        endDate = date.today() + timedelta(days=1)
        fromDate = endDate - timedelta(days=int(no_days_required))
        return fromDate, endDate

    @staticmethod
    def delete_file(file_path):
        # Check if file exists before deleting
        if os.path.exists(file_path):
            os.remove(file_path)
            print(f"File {file_path} has been deleted.")
        else:
            print(f"File {file_path} does not exist.")

    @staticmethod
    def check_previous_candle(previous_candle, current_candle, condition):
        return_value = 0
        if Configuration.candle_checking_conditions[0] == condition:
            '''previous_high_less_than_current_close'''
            if (previous_candle["High"] < current_candle["close"]):
                return_values = 1
        elif Configuration.candle_checking_conditions[1] == condition:
            ''''''
        return return_value

    @staticmethod
    def add_A_minus_B(df, columnA, columnB, new_col_name='diff_A_B'):
        """
        Adds a new column to the DataFrame that is the difference between:
        - columnA from the previous row
        - columnB from the current row

        Parameters:
            df (pd.DataFrame): The input DataFrame
            columnA (str): Column name to shift (previous row)
            columnB (str): Column name to subtract (current row)
            new_col_name (str): Name of the resulting column (optional)

        Returns:
            pd.DataFrame: DataFrame with the new column added
        """
        df[new_col_name] = df[columnA] - df[columnB]
        return df

    @staticmethod
    def add_shiftedA_minus_B(df, columnA, columnB, new_col_name='diff_prevA_currB'):
        """
        Adds a new column to the DataFrame that is the difference between:
        - columnA from the previous row
        - columnB from the current row

        Parameters:
            df (pd.DataFrame): The input DataFrame
            columnA (str): Column name to shift (previous row)
            columnB (str): Column name to subtract (current row)
            new_col_name (str): Name of the resulting column (optional)

        Returns:
            pd.DataFrame: DataFrame with the new column added
        """
        df[new_col_name] = df[columnA].shift(-1) - df[columnB]
        return df

    @staticmethod
    def add_A_minus_shiftedB(df, columnA, columnB, new_col_name='diff_prevA_currB'):
        """
        Adds a new column to the DataFrame that is the difference between:
        - columnA from the previous row
        - columnB from the current row

        Parameters:
            df (pd.DataFrame): The input DataFrame
            columnA (str): Column name to shift (previous row)
            columnB (str): Column name to subtract (current row)
            new_col_name (str): Name of the resulting column (optional)

        Returns:
            pd.DataFrame: DataFrame with the new column added
        """
        df[new_col_name] = df[columnA] - df[columnB].shift(-1)
        return df

    @staticmethod
    def add_candle_body_to_wick_ratio(df):
        """
        Adds a new column 'open_close_diff' to the DataFrame,
        which is the difference between 'open' and 'close'.

        Parameters:
            df (pd.DataFrame): Input DataFrame with 'open' and 'close' columns.

        Returns:
            pd.DataFrame: DataFrame with the new column added.
        """
        df['body'] = df['close'] - df['actual_open']
        df['body_wick_ratio'] = df['body'] / df['range'].replace(0, 0.001)

        return df

    @staticmethod
    def add_actual_open(df, open_col='open', close_col='close', new_col='actual_open'):
        """
        Adds a column with the 'actual open' value:
        - previous row's close if available
        - otherwise, the current row's open

        Parameters:
            df (pd.DataFrame): The input DataFrame with open and close columns.
            open_col (str): Name of the open column.
            close_col (str): Name of the close column.
            new_col (str): Name of the new column to be added.

        Returns:
            pd.DataFrame: The DataFrame with the new column added.
        """

        df[new_col] = df[close_col].shift(-1).fillna(df[open_col])
        return df

    @staticmethod
    def add_actual_low_high(df: pd.DataFrame):
        df['actual_high'] = df[['actual_open', 'high']].max(axis=1, skipna=True)
        df['actual_low'] =  df[['actual_open', 'low']].min(axis=1, skipna=True)
        return df

    @staticmethod
    def add_upper_wick_lower_wick(df):
        df['upper_wick'] = np.where(
            df['actual_open'] > df['close'],
            df['actual_high'] - df['actual_open'],
            df['actual_high'] - df['close']
        )

        df['lower_wick'] = np.where(
            df['actual_open'] > df['close'],
            df['close'] - df['actual_low'],
            df['actual_open'] - df['actual_low']
        )

        return df

    @staticmethod
    def add_upper_lower_wick_body_ratio(df):
        df['body_upper_wick_ratio'] = df['body'] / df['upper_wick'].replace(0, 0.001)
        df['body_lower_wick_ratio'] = df['body'] / df['lower_wick'].replace(0, 0.001)
        return df

    def reverse_rows(df):
        """
        Reverses the row order of the DataFrame (top to bottom).

        Parameters:
            df (pd.DataFrame): Input DataFrame.

        Returns:
            pd.DataFrame: DataFrame with reversed row order and reset index.
        """
        return df[::-1].reset_index(drop=True)


    @staticmethod
    def check_for_slow_down(reverse_df:pd.DataFrame):
        length = len(reverse_df)
        direction = 'direction'
        last_30_rows = reverse_df.tail(30)

        no_of_positive_candles = 0
        no_of_negative_candles = 0
        cumulative_positive_value = 0
        cumulative_negative_value = 0
        index = len(last_30_rows) - 1
        slow_down = ''
        if last_30_rows.iloc[index][direction] > 0:
            no_of_positive_candles = 1
            cumulative_positive_value = last_30_rows.iloc[index][direction]
            index -= 1
            last_index = index
            for i in range(index, 0, -1):
                direction_value = last_30_rows.iloc[i][direction]
                last_index -= 1
                if  direction_value > 0:
                    no_of_positive_candles += 1
                    cumulative_positive_value += direction_value
                else:
                    no_of_negative_candles += 1
                    cumulative_negative_value += direction_value

                    break
            for i in range(last_index, 0, -1):
                direction_value = last_30_rows.iloc[i][direction]

                if direction_value < 0:
                    no_of_negative_candles += 1
                    cumulative_negative_value += direction_value
                else:
                    cumulative_value = cumulative_negative_value + cumulative_positive_value
                    slow_down = str(no_of_negative_candles)+"_DOWN_"+str(no_of_positive_candles)+"_UP"+"_"+str(cumulative_value)
                    break

        else:
            no_of_negative_candles = 1
            cumulative_negative_value = last_30_rows.iloc[index][direction]
            index -= 1
            last_index = index
            for i in range(index, 0, -1):
                direction_value = last_30_rows.iloc[i][direction]
                last_index -= 1
                if direction_value < 0:
                    no_of_negative_candles += 1
                    cumulative_negative_value += direction_value
                else:
                    no_of_positive_candles += 1
                    cumulative_positive_value += direction_value

                    break
            for i in range(last_index, 0, -1):
                direction_value = last_30_rows.iloc[i][direction]

                if direction_value > 0:
                    no_of_positive_candles += 1
                    cumulative_positive_value += direction_value
                else:
                    cumulative_value = cumulative_negative_value + cumulative_positive_value
                    slow_down = str(no_of_positive_candles) + "_UP_" + str(
                        no_of_negative_candles) + "_DOWN_" + str(cumulative_value)
                    break
        return slow_down

    @staticmethod
    def convert_ist_to_utc_naive(dt_str: str) -> datetime:
        # Parse the datetime with IST offset
        print(dt_str)
        aware_dt = datetime.fromisoformat(dt_str)
        # Convert to UTC
        utc_dt = aware_dt.astimezone(ZoneInfo("UTC"))
        # Remove tzinfo so MySQL treats it as raw
        return utc_dt.replace(tzinfo=None)

    @staticmethod
    def to_naive_ist(dt_str: str) -> datetime:
        # Parse the datetime string including the +05:30 timezone
        aware_dt = datetime.fromisoformat(dt_str)  # Aware datetime with IST
        # Strip timezone info — retain the wall time (IST) as-is
        return aware_dt.replace(tzinfo=None)

    @staticmethod
    def check_for_multiple_slow_down(reverse_df:pd.DataFrame):

        SupportResistance().analyze_structure_breaks(reverse_df)
        length = len(reverse_df)
        print("Candles is given below ..... ")
        print(reverse_df.head(30))
        direction = 'direction'
        last_30_rows = reverse_df.tail(30)
        print(last_30_rows)
        no_of_positive_candles = 0
        no_of_negative_candles = 0
        cumulative_positive_value = 0
        cumulative_negative_value = 0
        index = len(last_30_rows) - 1
        slow_down = ''
        candles = []
        print("Slowdown === ")
        check_positive_candles = True
        if last_30_rows.iloc[index][direction] > 0:
            check_positive_candles = True
            no_of_positive_candles = 1
            cumulative_positive_value = last_30_rows.iloc[index][direction]

        else:
            check_positive_candles = False
            no_of_negative_candles = 1
            cumulative_negative_value = last_30_rows.iloc[index][direction]
        index -= 1

        for i in range(index, 0, -1):
            print(i)
            direction_value = last_30_rows.iloc[i][direction]
            if  direction_value > 0 and check_positive_candles == True:
                no_of_positive_candles += 1
                cumulative_positive_value += direction_value
            elif direction_value > 0 and check_positive_candles == False:
                check_positive_candles = True
                no_of_positive_candles = 1
                cumulative_positive_value = direction_value
                if no_of_negative_candles > 0:
                    string = str(no_of_negative_candles)+"_DN_"+str(round(cumulative_negative_value, 1))
                    candles.append(string)
            elif direction_value <= 0 and check_positive_candles == False:
                no_of_negative_candles += 1
                cumulative_negative_value += direction_value
            else :
                check_positive_candles = False
                no_of_negative_candles = 1
                cumulative_negative_value = direction_value
                if no_of_positive_candles > 0:
                    string = str(no_of_positive_candles)+'_UP_'+ str(round(cumulative_positive_value, 1))
                    candles.append(string)

        if check_positive_candles == True:
            candles.append(str(no_of_positive_candles)+'_UP_'+ str(round(cumulative_positive_value, 1)))
        else:
            candles.append(str(no_of_negative_candles)+"_DN_"+str(round(cumulative_negative_value, 1)))
        candles.reverse()
        for i in candles:
            print(i)
        return candles


    @staticmethod
    def check_for_breach(overall_candle: dict, index: int, reverse_df:pd.DataFrame):
        '''
        Check if there is a breach based on the overall_candle and the current candle
        :param overall_candle: consolidated candle based on the territory
        :param current_candle: current candle
        :return: true if breach happened else false
        '''
        current_candle = reverse_df.loc[reverse_df["index"] == index].iloc[0]
        if current_candle['direction'] > 0:
            '''
            Positive Candle. Check if this is broken on Buyers side
            '''
            if current_candle['close'] > overall_candle['actual_high']:
                # Breach happened. Set the values of overall_candle based on the breach
                overall_candle['close'] = current_candle['close']
                overall_candle['actual_high'] = current_candle['actual_high']

                # Find the Guy who started the move and assign the overall_candle["open"] and low
                # We have to go through the candles up until overall_candle["breach_candle_index"]
                for i in range(current_candle['index'], overall_candle["guy_who_started_index"] - 1, -1):
                    row = reverse_df.loc[reverse_df['index'] == i].iloc[0]
                    # Since we have breached on the buyers side - Check if this is a pull back candle -  direction of this candle < 0
                    if row['direction'] < 0:
                        # This is a pull back candle
                        overall_candle['actual_open'] = row['close']
                        overall_candle['actual_low'] = row['actual_low']
                        overall_candle['guy_who_started_index'] = row['index']
                        # Do we need to go and find further below?
                        # print(current_candle['datetime'])
                        # print(overall_candle)
                        break
        elif current_candle['direction'] < 0:
            '''
            Negative Candle. Check if this has broken on Sellers side
            '''
            if current_candle['close'] < overall_candle['actual_low']:
                # Breach happened. Set the values of overall_candle based on the breach
                overall_candle['close'] = current_candle['close']
                overall_candle['actual_low'] = current_candle['actual_low']

                # Find the Guy who started the move and assign the overall_candle["open"] and high
                # We have to go through the candles up until overall_candle["breach_candle_index"]
                for i in range(current_candle['index'], overall_candle["guy_who_started_index"] - 1, -1):
                    row = reverse_df.loc[reverse_df['index'] == i].iloc[0]
                    # Since we have breached on the sellers side - Check if this is a pull back candle -  direction of this candle > 0
                    if row['direction'] > 0:
                        # This is a pull back candle
                        overall_candle['actual_open'] = row['close']
                        overall_candle['actual_high'] = row['actual_high']
                        overall_candle['guy_who_started_index'] = row['index']
                        # Do we need to go and find further below?
                        #print(row['datetime'])
                        #print(overall_candle)
                        break




    @staticmethod
    def find_sellers_buyers_territory(df):
        territory = []
        open = "actual_open"
        high = "actual_high"
        low = "actual_low"
        close = "close"
        guy_who_started_index = "guy_who_started_index"
        guy_who_started_date = "guy_who_started_date"
        reverse_df = Utility.reverse_rows(df)
        reverse_df['index'] = range(len(reverse_df))
        first_record = reverse_df.iloc[0]
        slow_down = Utility.check_for_slow_down(reverse_df)
        Utility.check_for_multiple_slow_down(reverse_df)
        print("\n\nSlowDown == "+slow_down+"\n\n")
        #print(reverse_df.head(30))
        # print(first_record)
        # print(type(first_record))

        if first_record["direction"] > 0:
            # The candle is green
            territory = {
                open: first_record[open],
                close: first_record[close],
                high: first_record[high],
                low: first_record[low],
                guy_who_started_index: 0
            }
        else:
            # The candle is red
            territory = {

                open: first_record[open],
                close: first_record[close],
                high: first_record[high],
                low: first_record[low],
                guy_who_started_index: 0
            }

        #print(territory)
        for i in range(1, len(reverse_df)):
            Utility.check_for_breach(territory, i, reverse_df)
        print("Final ")
        territory[guy_who_started_date] = reverse_df[reverse_df['index'] == territory['guy_who_started_index']].iloc[0]['datetime']
        territory['territory_value'] = territory['close'] - territory['actual_open']
        if territory['territory_value'] > 0:
            territory['territory'] = 'Buyers'
        else:
            territory['territory'] = 'Sellers'
        # print(territory)
        # print(reverse_df[reverse_df['index'] == territory['guy_who_started_index']].iloc[0]['datetime'])
        return territory

